import os
import openpyxl

def sort_excel_by_first_column(file_path, sheet_name):
    # Load the workbook and sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Read all rows from the sheet
    rows = list(sheet.iter_rows(values_only=True))
    
    # Separate header from data rows (assume the first row is the header)
    header = rows[0]
    data_rows = rows[1:]
    
    # Sort rows by the first column (alphabetically)
    sorted_rows = sorted(data_rows, key=lambda row: str(row[0]))
    
    # Clear the sheet and write the sorted data back
    sheet.delete_rows(1, sheet.max_row)  # Clear the existing data
    sheet.append(header)  # Add the header back
    for row in sorted_rows:
        sheet.append(row)
    
    # Save the changes to the same file (overwrite the original)
    workbook.save(file_path)
    print(f"Excel file sorted and saved back to {file_path}")

def sort_excel_files_in_folder(folder_path, sheet_name):
    # Loop through all files in the folder
    for file_name in os.listdir(folder_path):
        # Check if the file is an Excel file (ending with .xlsx)
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file_name)
            print(f"Processing {file_path}...")
            sort_excel_by_first_column(file_path, sheet_name)

# Example usage
folder_path = r"C:\Users\Ayush Srivastava\Desktop\task3\main\undergrad" # Replace with the path to your folder containing Excel files
sheet_name = "UserIDInfo"          # Replace with your sheet name

sort_excel_files_in_folder(folder_path, sheet_name)


